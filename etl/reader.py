"""원본 실적 파일 → 파이썬 dict 로 읽어오는 계층.

지원 포맷
  * .xlsb  (pyxlsb)          — 26.8월 이후 '설계사시상 진척현황'
  * .xlsx  (openpyxl)        — 그 이전 '설계사 인별실적시트'
                               암호가 걸려 있으면 msoffcrypto 로 먼저 해제

두 포맷 모두 '설계사' 시트의 컬럼 레이아웃이 동일해서 하나의 파서로 처리한다.
파싱 결과는 etl/cache/ 에 JSON 으로 캐시한다(파일 크기+mtime 이 키).
"""
from __future__ import annotations

import datetime
import io
import json
import os
import re
from pathlib import Path

import schema

PASSWORDS = ["0000", ""]

EPOCH = datetime.date(1899, 12, 30)


def excel_date(v):
    """엑셀 시리얼 / datetime / 문자열 → datetime.date"""
    if v is None:
        return None
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    if isinstance(v, (int, float)):
        return EPOCH + datetime.timedelta(days=int(v))
    m = re.search(r"(\d{4})[-./](\d{1,2})[-./](\d{1,2})", str(v))
    if m:
        return datetime.date(int(m[1]), int(m[2]), int(m[3]))
    return None


# --------------------------------------------------------------------------
# 암호 해제 (두 포맷 공통)
# --------------------------------------------------------------------------
def _opened(path: Path):
    """암호가 걸려 있으면 해제해서 BytesIO 로, 아니면 그대로 파일 핸들을 연다.

    같은 '설계사시상 진척현황' 이름이라도 어느 날은 평문 .xlsb, 어느 날은
    암호 걸린 .xlsb(구형 .xlsx 와 같은 OLE2 CFB 컨테이너)로 온다.
    확장자로 판단하지 않고 msoffcrypto 의 is_encrypted() 로 실제 상태를 본다.
    """
    import msoffcrypto

    fh = open(path, "rb")
    try:
        office = msoffcrypto.OfficeFile(fh)
        if not office.is_encrypted():
            fh.seek(0)
            return fh
    except Exception:
        # 암호 여부 판별 자체가 실패하면(=이 포맷이 아니면) 원본 그대로 읽는다.
        fh.seek(0)
        return fh

    for pw in PASSWORDS:
        try:
            office.load_key(password=pw)
            buf = io.BytesIO()
            office.decrypt(buf)
            buf.seek(0)
            fh.close()
            return buf
        except Exception:
            continue
    fh.close()
    raise RuntimeError(f"암호 해제 실패: {path.name}")


# --------------------------------------------------------------------------
# 포맷별 raw 행 읽기 : [{col_index: value}, ...]
# --------------------------------------------------------------------------
def _rows_xlsb(path: Path, sheet: str):
    from pyxlsb import open_workbook

    src = _opened(path)
    wb = open_workbook(src)
    if sheet not in wb.sheets:
        return None
    out = []
    with wb.get_sheet(sheet) as sh:
        for row in sh.rows():
            out.append({c.c: c.v for c in row if c.v is not None})
    return out


def _rows_xlsx(path: Path, sheet: str):
    from openpyxl import load_workbook

    src = _opened(path)
    wb = load_workbook(src, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return None
    ws = wb[sheet]
    out = []
    for row in ws.iter_rows(values_only=True):
        out.append({i: v for i, v in enumerate(row) if v is not None})
    wb.close()
    return out


def raw_rows(path: Path, sheet: str):
    if path.suffix.lower() == ".xlsb":
        return _rows_xlsb(path, sheet)
    return _rows_xlsx(path, sheet)


# --------------------------------------------------------------------------
# '설계사' 시트 → 정규화 레코드
# --------------------------------------------------------------------------
def num(v):
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def parse_snapshot(path: Path):
    """한 파일 → {'as_of': 'YYYY-MM-DD', 'people': [...], 'branches': [...]}"""
    rows = raw_rows(path, "설계사")
    if not rows:
        raise RuntimeError(f"'설계사' 시트를 찾을 수 없음: {path.name}")

    as_of = excel_date(rows[2].get(0)) if len(rows) > 2 else None
    if as_of is None:  # 파일명 260814_... 로 폴백
        m = re.match(r"(\d{2})(\d{2})(\d{2})", path.name)
        as_of = datetime.date(2000 + int(m[1]), int(m[2]), int(m[3])) if m else None

    warnings = []
    header = rows[7] if len(rows) > 7 else {}
    for idx, text in schema.HEADER_CHECK:
        got = header.get(idx)
        if got != text:
            warnings.append(f"헤더 불일치 col{idx}: 기대 '{text}' / 실제 '{got}'")

    C = schema.COL
    people = []
    for r in rows[8:]:
        if not r.get(C["person"]):
            continue
        rec = {
            "manager": r.get(C["manager"]),
            "hq": r.get(C["hq"]),
            "dept": r.get(C["dept"]),
            "deptCode": int(num(r.get(C["dept_code"]))) or None,
            "center": r.get(C["center"]),
            "centerCode": int(num(r.get(C["center_code"]))) or None,
            "centerHead": r.get(C["center_head"]),
            "agency": r.get(C["agency"]),
            "branch": r.get(C["branch"]),
            "branchCode": int(num(r.get(C["branch_code"]))) or None,
            "name": r.get(C["person"]),
            "code": int(num(r.get(C["person_code"]))) or None,
            "convP": round(num(r.get(C["conv_p"])), 3),
            "cred": round(num(r.get(C["cred_month"])), 3),
            "weeks": [round(num(r.get(c)), 3) for c in schema.WEEK_COLS],
        }
        awards = {}
        for key, _label, flag_col, perf_cols, money_col in schema.AWARDS:
            flag = num(r.get(flag_col))
            if not flag:
                continue
            awards[key] = {
                "perf": round(sum(num(r.get(c)) for c in perf_cols), 3),
                "money": round(num(r.get(money_col)), 1),
            }
        rec["awards"] = awards
        rec["prestige"] = {
            "on": bool(num(r.get(schema.PRESTIGE["flag"]))),
            "m7": round(num(r.get(schema.PRESTIGE["m7"])), 3),
            "m8": round(num(r.get(schema.PRESTIGE["m8"])), 3),
            "grade": num(r.get(schema.PRESTIGE["grade"])),
        }
        people.append(rec)

    # '지사' 시트: 지사별 재적 설계사수 (xlsb 에만 존재)
    branches = []
    try:
        brows = raw_rows(path, "지사")
    except Exception:
        brows = None
    if brows:
        B = schema.BRANCH_COL
        for r in brows[schema.BRANCH_DATA_START:]:
            if not r.get(B["branch"]):
                continue
            branches.append({
                "hq": r.get(B["hq"]),
                "dept": r.get(B["dept"]),
                "deptCode": int(num(r.get(B["dept_code"]))) or None,
                "center": r.get(B["center"]),
                "centerCode": int(num(r.get(B["center_code"]))) or None,
                "centerHead": r.get(B["center_head"]),
                "agency": r.get(B["agency"]),
                "branch": r.get(B["branch"]),
                "branchCode": int(num(r.get(B["branch_code"]))) or None,
                "headcount": int(num(r.get(B["headcount"]))),
                "worldTourOperating": int(num(r.get(B["wt_operating"]))),
                "worldTour": [
                    {
                        "key": key, "label": label,
                        "people": int(num(r.get(cnt_col))),
                        "reached": bool(num(r.get(tier_col))),
                    }
                    for key, label, cnt_col, tier_col in schema.WORLD_TOUR_TIERS
                ],
            })

    return {
        "asOf": as_of.isoformat() if as_of else None,
        "source": path.name,
        "warnings": warnings,
        "people": people,
        "branches": branches,
    }


# --------------------------------------------------------------------------
# 캐시
# --------------------------------------------------------------------------
def load_snapshot(path: Path, cache_dir: Path):
    cache_dir.mkdir(parents=True, exist_ok=True)
    st = path.stat()
    key = f"{path.stem}__{st.st_size}_{int(st.st_mtime)}.json"
    cached = cache_dir / key
    if cached.exists():
        return json.loads(cached.read_text(encoding="utf-8"))
    snap = parse_snapshot(path)
    cached.write_text(json.dumps(snap, ensure_ascii=False), encoding="utf-8")
    # 같은 파일의 옛 캐시 정리
    for old in cache_dir.glob(f"{path.stem}__*.json"):
        if old.name != key:
            old.unlink(missing_ok=True)
    return snap


SNAPSHOT_PATTERNS = ("*설계사시상 진척현황*", "*설계사 인별실적시트*")


def discover(src_dir: Path):
    """소스 폴더에서 스냅샷 후보 파일 수집. 같은 기준일이면 최신 mtime 1개만."""
    seen = {}
    for pat in SNAPSHOT_PATTERNS:
        for p in src_dir.glob(pat):
            if p.suffix.lower() not in (".xlsb", ".xlsx"):
                continue
            if p.name.startswith("~$"):
                continue
            m = re.match(r"(\d{6})", p.name)
            stamp = m[1] if m else p.stem
            prev = seen.get(stamp)
            if prev is None or p.stat().st_mtime > prev.stat().st_mtime:
                seen[stamp] = p
    return [seen[k] for k in sorted(seen)]
