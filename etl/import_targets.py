"""'매출진척.xlsx' (표지 시트) → etl/targets.json

원본 한 장에 세 가지가 같이 들어 있다.
  · 월간 매출목표
  · 전월 10만가동 인원 (실적)
  · 20만가동 목표 인원

행이 [계 → 팀 → 영업단 → 센터] 로 섞여 있고 **레벨마다 단위가 다르다**.
센터는 천원, 영업단·팀은 백만원. 그래서 센터 행만 읽고 상위는 합산으로 만든다.
(합산이 파일의 상위 행과 일치하는지 검증한 뒤 통과시킨다.)

사용법
    python3 etl/import_targets.py '/path/매출진척.xlsx' --month 2026-08
"""
from __future__ import annotations

import argparse
import datetime
import json
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

import reader  # noqa: E402

HERE = Path(__file__).resolve().parent
TARGETS = HERE / "targets.json"

COL = {"dept": 2, "center": 3, "head": 4, "revenue": 5, "prev10man": 6, "target20man": 7}
DEPT_SUFFIX = "컨설팅영업단"


def load_rows(path: Path):
    from openpyxl import load_workbook

    wb = load_workbook(reader._opened(path), read_only=True, data_only=True)
    ws = wb["표지"] if "표지" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def parse(path: Path):
    """센터 행을 먼저 읽고, 그 센터들이 속한 영업단 행만 검증용으로 집는다.

    같은 표에 [計 → 팀 → 영업단 → 센터] 가 섞여 있는데 이름만으로는
    팀과 영업단을 가르기 어렵다(충청호남·부산대구는 팀인데 '팀'으로 안 끝난다).
    센터가 실제로 딸린 이름만 영업단으로 인정하면 그 모호함이 사라진다.
    """
    rows = load_rows(path)
    centers, dept_rows = {}, {}

    for r in rows:
        if len(r) <= COL["target20man"]:
            continue
        d, c = r[COL["dept"]], r[COL["center"]]
        if not isinstance(d, str) or not d.strip():
            continue
        d = d.strip()
        rev = num(r[COL["revenue"]])
        if rev is None:
            continue

        if isinstance(c, str) and c.strip():
            # 센터 행 — 단위 천원
            centers[f"{d}.{c}".replace(" ", "")] = {
                "revenue": int(rev),
                "prev10man": int(num(r[COL["prev10man"]]) or 0),
                "target20man": int(num(r[COL["target20man"]]) or 0),
                "head": (r[COL["head"]] or "").strip() or None,
                "dept": f"{d}{DEPT_SUFFIX}",
            }
        else:
            # 상위 행(계·팀·영업단) — 단위 백만원. 어느 쪽인지는 아래에서 가른다.
            dept_rows[d] = {
                "revenueM": rev,
                "prev10man": int(num(r[COL["prev10man"]]) or 0),
                "target20man": int(num(r[COL["target20man"]]) or 0),
                "head": (r[COL["head"]] or "").strip() or None,
            }

    have = {c["dept"] for c in centers.values()}
    depts = {f"{k}{DEPT_SUFFIX}": v for k, v in dept_rows.items()
             if f"{k}{DEPT_SUFFIX}" in have}
    return centers, depts


def verify(centers, depts):
    """센터 합이 파일의 영업단 행과 맞는지 확인. 안 맞으면 레이아웃이 바뀐 것."""
    problems = []
    for dname, d in depts.items():
        cs = [c for c in centers.values() if c["dept"] == dname]
        if not cs:
            problems.append(f"{dname}: 소속 센터를 못 찾음")
            continue
        s_rev = sum(c["revenue"] for c in cs) / 1000
        if abs(s_rev - d["revenueM"]) > 0.05:
            problems.append(f"{dname}: 매출목표 합 {s_rev:.2f} ≠ 파일 {d['revenueM']:.2f} (백만원)")
        for k in ("prev10man", "target20man"):
            s = sum(c[k] for c in cs)
            if s != d[k]:
                problems.append(f"{dname}: {k} 합 {s} ≠ 파일 {d[k]}")
    return problems


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("src", type=Path)
    ap.add_argument("--month", required=True, help="예: 2026-08")
    a = ap.parse_args()

    centers, depts = parse(a.src)
    if not centers:
        raise SystemExit("센터 행을 하나도 못 읽었습니다. 시트 레이아웃을 확인하세요.")

    problems = verify(centers, depts)
    if problems:
        print("✗ 검증 실패 — 상위 행 합계가 안 맞습니다:")
        for p in problems:
            print("   ", p)
        raise SystemExit(1)

    doc = json.loads(TARGETS.read_text(encoding="utf-8")) if TARGETS.exists() else {}
    doc.setdefault("_설명", "월간 매출목표(천원) · 전월 10만가동(명) · 20만가동 목표(명)")
    doc.setdefault("_주의", "etl/import_targets.py 로 생성한다. 손으로 고치지 말 것.")
    doc.setdefault("unit", "천원")
    doc.setdefault("months", {})
    doc["months"][a.month] = {
        "source": a.src.name,
        "importedAt": datetime.datetime.now().isoformat(timespec="seconds"),
        "deptHeads": {k: v["head"] for k, v in depts.items() if v["head"]},
        "centers": {k: {kk: vv for kk, vv in v.items() if kk != "dept"}
                    for k, v in sorted(centers.items())},
    }
    TARGETS.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")

    tot = sum(c["revenue"] for c in centers.values())
    print(f"✓ {a.month} — 센터 {len(centers)}개 / 영업단 {len(depts)}개")
    print(f"   매출목표 합 {tot:,}천원 ({tot/10:,.0f}만원)")
    print(f"   전월 10만가동 {sum(c['prev10man'] for c in centers.values()):,}명 · "
          f"20만가동 목표 {sum(c['target20man'] for c in centers.values()):,}명")
    print(f"   → {TARGETS}")


if __name__ == "__main__":
    main()
